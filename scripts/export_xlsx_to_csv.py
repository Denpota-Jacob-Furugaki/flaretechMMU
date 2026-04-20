"""Export every tab of every .xlsx under src/ (or root, as fallback) to data/raw/<workbook>/.

Canonical location for source workbooks is `src/`. Files in the project root are
still picked up for backward compatibility, but if the same filename exists in
both places, the `src/` copy wins.

Skipped:
- Excel lock files starting with `~$`
- Duplicate-download suffixes like `... (1).xlsx`, `... (2).xlsx`

Run from project root:  python scripts/export_xlsx_to_csv.py
"""
from __future__ import annotations
import csv
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
RAW = ROOT / "data" / "raw"

DUPE_SUFFIX_RE = re.compile(r" \(\d+\)$")


def slugify(name: str) -> str:
    s = name.strip().replace(" ", "_")
    s = re.sub(r"[\\/:*?\"<>|]", "_", s)
    return s or "untitled"


def export_workbook(xlsx_path: Path) -> list[tuple[str, int, int]]:
    out_dir = RAW / slugify(xlsx_path.stem)
    out_dir.mkdir(parents=True, exist_ok=True)
    results: list[tuple[str, int, int]] = []

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_path = out_dir / f"{slugify(sheet_name)}.csv"
        rows_written = 0
        max_cols = 0
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            trailing_blank = 0
            buffer: list[list] = []
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else v for v in row]
                while values and values[-1] == "":
                    values.pop()
                if not values:
                    trailing_blank += 1
                    buffer.append([])
                    continue
                if trailing_blank:
                    for _ in buffer:
                        writer.writerow([])
                    buffer = []
                    trailing_blank = 0
                writer.writerow(values)
                rows_written += 1
                max_cols = max(max_cols, len(values))
        results.append((sheet_name, rows_written, max_cols))
    wb.close()
    return results


def discover() -> list[Path]:
    """Find xlsx files to process. `src/` wins over root on name collision."""
    seen: dict[str, Path] = {}
    # root first, src/ second, so src/ overrides
    for base in (ROOT, SRC):
        if not base.exists():
            continue
        for xlsx in sorted(base.glob("*.xlsx")):
            if xlsx.name.startswith("~$"):
                continue
            if DUPE_SUFFIX_RE.search(xlsx.stem):
                continue
            seen[xlsx.name] = xlsx
    return sorted(seen.values(), key=lambda p: p.name)


def main() -> int:
    xlsx_files = discover()
    if not xlsx_files:
        print("no .xlsx files found in", ROOT, "or", SRC)
        return 1

    for xlsx in xlsx_files:
        rel = xlsx.relative_to(ROOT)
        print(f"\n== {rel} ==")
        for sheet_name, rows, cols in export_workbook(xlsx):
            print(f"  {sheet_name:40s}  rows={rows:5d}  cols={cols:3d}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
